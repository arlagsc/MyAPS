from flask import Flask, render_template, request, jsonify, redirect, url_for
from database import init_db, reset_db, get_db_connection
from scheduler import run_advanced_scheduling
from scheduler_core import GreedyScheduler, GeneticScheduler, SimulatedAnnealingScheduler
from api.routes import api_bp
from api.mes_orders import mes_api_bp
import datetime

app = Flask(__name__)
init_db()

app.register_blueprint(api_bp)
app.register_blueprint(mes_api_bp)

@app.route('/')
def index():
    conn = get_db_connection()
    resources = conn.execute("SELECT * FROM resources").fetchall()
    conn.close()
    return render_template('index.html', resources=resources)

@app.route('/manage')
def manage():
    conn = get_db_connection()
    resources = conn.execute("SELECT * FROM resources").fetchall()
    orders = conn.execute("""
        SELECT * FROM work_orders 
        ORDER BY plan_type DESC, priority ASC
    """).fetchall()
    conn.close()
    return render_template('manage.html', resources=resources, orders=orders)

@app.route('/api/schedule_data')
def get_gantt_data():
    conn = get_db_connection()
    
    # 1. 获取资源组
    resources = conn.execute("SELECT * FROM resources ORDER BY id").fetchall()
    groups = []
    for r in resources:
        groups.append({
            "id": r['id'],
            "content": f"<span style='font-weight:bold;'>{r['name']}</span>",
            "value": r['id']
        })

    # 2. 获取任务数据
    orders = conn.execute("""
        SELECT w.*, r.name as res_name 
        FROM work_orders w
        JOIN resources r ON w.resource_id = r.id
        WHERE w.status IN ('Scheduled', 'Delayed', 'Warning-Debounce') 
           OR w.status LIKE 'Scheduled%' 
           OR w.status LIKE 'Delayed%'
    """).fetchall()
    
    # --- [阶段四核心] 预处理：记录所有 B 面的结束时间，用于校验 48 小时规则 ---
    b_side_ends = {}
    for row in orders:
        if row['smt_side'] in ['B', 'B面', 'Back'] and row['planned_end']:
            try:
                e_str = str(row['planned_end']).replace('T', ' ').replace('/', '-').strip()
                e_dt = datetime.datetime.strptime(e_str, '%Y-%m-%d %H:%M') if len(e_str) <= 16 else datetime.datetime.strptime(e_str, '%Y-%m-%d %H:%M:%S')
                b_side_ends[row['job_id']] = e_dt
            except:
                pass

    items = []
    
    # --- [阶段四核心] 渲染设备保养阴影块 ---
    try:
        maints = conn.execute("SELECT * FROM equipment_maintenance").fetchall()
        for m in maints:
            items.append({
                "id": f"maint_{m['id']}",
                "group": m['line_id'],
                "start": m['start_time'].replace(' ', 'T'),
                "end": m['end_time'].replace(' ', 'T'),
                "type": "background",
                "className": "vis-item-maintenance",
                "title": f"【{m['type']}】 {m['reason']}"
            })
    except Exception as e:
        print(f"读取保养记录失败: {e}")

    conn.close()

    for row in orders:
        is_time_insufficient = False
        is_ab_violation = False
        duration_min = 0
        std_time = float(row['std_time']) if row['std_time'] else 0
        
        s_dt, e_dt = None, None
        
        if row['planned_start'] and row['planned_end']:
            try:
                s_str = str(row['planned_start']).replace('T', ' ').replace('/', '-').strip()
                e_str = str(row['planned_end']).replace('T', ' ').replace('/', '-').strip()
                
                s_dt = datetime.datetime.strptime(s_str, '%Y-%m-%d %H:%M') if len(s_str) <= 16 else datetime.datetime.strptime(s_str, '%Y-%m-%d %H:%M:%S')
                e_dt = datetime.datetime.strptime(e_str, '%Y-%m-%d %H:%M') if len(e_str) <= 16 else datetime.datetime.strptime(e_str, '%Y-%m-%d %H:%M:%S')
                
                duration_min = (e_dt - s_dt).total_seconds() / 60
                
                if std_time > 0 and duration_min < (std_time - 1):
                    is_time_insufficient = True
                    
                # --- [阶段四核心] 校验 A 面是否超过 B 面产出 48 小时 ---
                if row['smt_side'] in ['A', 'A面', 'Front'] and row['job_id'] in b_side_ends:
                    b_end = b_side_ends[row['job_id']]
                    gap_hours = (s_dt - b_end).total_seconds() / 3600
                    if gap_hours > 48:
                        is_ab_violation = True
            except:
                pass 

        className = 'vis-item-blue'
        content_html = row['task_id']

        # 样式优先级判定 (报警优先)
        if is_ab_violation:
            className = 'vis-item-error'
            content_html = f"🚨 {row['task_id']} <small>(>48h违反!)</small>"
        elif is_time_insufficient:
            className = 'vis-item-warning'
            content_html = f"⚠️ {row['task_id']} <small>({int(duration_min)}/{int(std_time)}m)</small>"
        # ===== [新增] 防抖违规警报 =====
        elif row['status'] == 'Warning-Debounce':
            className = 'vis-item-debounce'
            content_html = f"🔄 {row['task_id']} <small>(防抖超限)</small>"
        # ==============================
        elif row['status'].startswith('Delayed') or '⚠️' in row['status']:
            className = 'vis-item-red'
        elif row['is_locked']:
            className = 'vis-item-dark'
        elif row['plan_type'] == 'SIMULATION':
            className = 'vis-item-striped'

        items.append({
            "id": row['task_id'],
            "group": row['resource_id'],
            "content": content_html, 
            "start": row['planned_start'].replace(' ', 'T'),
            "end": row['planned_end'].replace(' ', 'T'),
            "className": className,
            "title": f"任务: {row['task_id']}<br>标准: {int(std_time)}m | 计划: {int(duration_min)}m",
            "data": {
                "name": row['task_id'],
                "resource_id": row['resource_id'],
                "is_locked": row['is_locked'],
                "std_time": std_time, 
                "status": row['status']
            }
        })
        
    return jsonify({"groups": groups, "items": items})

@app.route('/api/run', methods=['POST'])
def run_schedule():
    data = request.json
    mode = data.get('mode', 'SIMULATION')
    algo = data.get('algorithm', 'greedy')
    
    conn = get_db_connection()
    try:
        orders_db = conn.execute("SELECT * FROM work_orders WHERE status IN ('Pending', 'Scheduled') AND (is_locked = 0 OR is_locked IS NULL)").fetchall()
        resources_db = conn.execute("SELECT * FROM resources").fetchall()
        
        orders_list = [dict(row) for row in orders_db]
        resources_list = [dict(row) for row in resources_db]

        if not orders_list:
             return jsonify({'success': True, 'message': '没有可排产的任务'})
        
        scheduler = None
        if algo == 'ga':
            #scheduler = GeneticScheduler(orders_list, resources_list)
            #best_schedule = scheduler.run(pop_size=50, generations=30)
            # 激活真正的 AI：遗传算法！
            #from scheduler_core import GeneticScheduler
            scheduler = GeneticScheduler(orders_list, resources_list)
            # pop_size=15, generations=10 约等于在后台模拟推演了 150 次全局排产
            best_schedule = scheduler.run(pop_size=15, generations=10)
        elif algo == 'sa':
            scheduler = SimulatedAnnealingScheduler(orders_list, resources_list)
            best_schedule = scheduler.run(initial_temp=10000)
        else:
            scheduler = GreedyScheduler(orders_list, resources_list)
            best_schedule = scheduler.run()

        for item in best_schedule:
            s_str = item['planned_start'].strftime('%Y-%m-%d %H:%M')
            e_str = item['planned_end'].strftime('%Y-%m-%d %H:%M')
            
            # 【防抖违规捕获】：如果打破了 3天/7天 的约束，标记为专属状态
            status = 'Warning-Debounce' if item.get('violated') else 'Scheduled'
            
            conn.execute("""
                UPDATE work_orders 
                SET planned_start = ?, planned_end = ?, status = ?, plan_type = ?, resource_id = ?
                WHERE task_id = ?
            """, (s_str, e_str, status, mode, item['resource_id'], item['task_id']))

        conn.commit()
        return jsonify({'success': True, 'message': f'排产完成！已调度 {len(best_schedule)} 个任务'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})
    finally:
        conn.close()

@app.route('/api/update_order_manual', methods=['POST'])
def update_order_manual():
    data = request.json
    task_id = data.get('task_id')
    start_str = data.get('planned_start')
    end_str = data.get('planned_end')
    is_locked = data.get('is_locked')
    resource_id = data.get('resource_id')

    if start_str: start_str = start_str.replace('T', ' ')
    if end_str: end_str = end_str.replace('T', ' ')

    conn = get_db_connection()
    new_status = 'Scheduled' if (start_str and end_str) else 'Pending'
    
    if resource_id:
        conn.execute('''UPDATE work_orders SET planned_start = ?, planned_end = ?, is_locked = ?, status = ?, resource_id = ? WHERE task_id = ?''', (start_str, end_str, is_locked, new_status, resource_id, task_id))
    else:
        conn.execute('''UPDATE work_orders SET planned_start = ?, planned_end = ?, is_locked = ?, status = ? WHERE task_id = ?''', (start_str, end_str, is_locked, new_status, task_id))
    
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": f"任务 {task_id} 已更新"})

@app.route('/manage/reset', methods=['GET', 'POST'])
def reset_data():
    """重置系统：清空旧数据，并从接口获取 MES 已经拆分好的工单数据"""
    from database import get_db_connection
    from datetime import datetime
    from flask import redirect, url_for
    import traceback
    
    try:
        # 1. 清空旧工单
        conn = get_db_connection()
        conn.execute('DELETE FROM work_orders')
        conn.commit()
        
        # 2. 从模拟 MES/SAP 接口同步已拆分好的工序级工单
        from adapters.base import AdapterFactory
        sap = AdapterFactory.get_sap_adapter()
        orders = sap.get_orders_from_zpp008() # 现在拿到的是直接可以排产的子任务
        now = datetime.now().isoformat()
        
        saved = 0
        for order in orders:
            # 3. 直接将接口传来的明细无脑入库，彻底实现系统解耦
            conn.execute("""
                INSERT INTO work_orders (
                    task_id, job_id, product_code, resource_id, qty, std_time,
                    priority, material_time, software_time, deadline, smt_side,
                    related_task_id, process_req, status, planned_start, planned_end,
                    setup_time, is_locked, plan_type, created_at, updated_at,
                    workshop, component_code, component_desc
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                order.get('task_id'), order.get('job_id'), order.get('product_code'), None, 
                order.get('qty'), order.get('std_time'), order.get('priority'), 
                '', '', order.get('demand_date'), order.get('smt_side'), 
                order.get('related_task_id'), None, 'Pending', None, None, 
                10, 0, 'OFFICIAL', now, now,
                order.get('workshop'), order.get('component_code'), order.get('component_desc')
            ))
            saved += 1
        
        conn.commit()
        conn.close()
        return redirect(url_for('config_page') + '?reset=ok&count=' + str(saved))
        
    except Exception as e:
        traceback.print_exc()
        return redirect(url_for('config_page') + '?reset=fail&msg=' + str(e))
    
# --- [阶段四核心] 新增设备保养 API ---
@app.route('/api/maintenance', methods=['POST'])
def add_maintenance():
    data = request.json
    conn = get_db_connection()
    try:
        conn.execute('''
            INSERT INTO equipment_maintenance (line_id, start_time, end_time, type, reason)
            VALUES (?, ?, ?, ?, ?)
        ''', (data['line_id'], data['start_time'].replace('T', ' '), data['end_time'].replace('T', ' '), data['type'], data['reason']))
        conn.commit()
        return jsonify({"success": True, "message": "已添加设备停机时间！"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})
    finally:
        conn.close()

# 其余页面路由保持极简结构（为节省篇幅省略未修改部分，确保核心调度 API 完整）
@app.route('/resources')
def resources_page(): return render_template('resources.html')
@app.route('/config')
def config_page(): return render_template('config.html')
@app.route('/dashboard/<workshop>')
def dashboard_workshop(workshop): return render_template('dashboard_workshop.html', workshop_code=workshop.upper(), workshop_name=workshop.upper())

# === 🛠️ 终极修复工具：同步真实的 SMT/DIP 产线基础数据 ===
@app.route('/fix_resources')
def fix_resources():
    conn = get_db_connection()
    try:
        # 1. 清空旧的无效演示资源
        conn.execute("DELETE FROM resources")
        
        # 2. 按照 PRD 生成所有真实的 SMT 线体
        smt_lines = [f'S{i:02d}' for i in list(range(1, 50)) + [98, 99]]
        for line in smt_lines:
            conn.execute("INSERT INTO resources (id, name, type) VALUES (?, ?, ?)", 
                         (line, f'SMT线-{line}', 'SMT'))
            
        # 3. 按照 PRD 生成所有真实的 DIP 线体
        dip_lines = [f'D{i:02d}' for i in range(1, 23)]
        for line in dip_lines:
            conn.execute("INSERT INTO resources (id, name, type) VALUES (?, ?, ?)", 
                         (line, f'DIP线-{line}', 'DIP'))
            
        conn.commit()
        return "<h3>✅ 产线资源修复成功！已自动生成真实的 S01-S49 及 D01-D22 线体。</h3><a href='/manage'>👉 点击这里返回数据管理中心</a>"
    except Exception as e:
        return f"修复失败: {e}"
    finally:
        conn.close()

# === 新增：车间日排产计划矩阵 API (类 Excel 视图) ===
@app.route('/api/schedule_matrix/<workshop>')
def schedule_matrix(workshop):
    from datetime import datetime, timedelta
    conn = get_db_connection()
    
    # 筛选该车间已排产的任务
    if workshop == 'ALL':
        orders = conn.execute("SELECT * FROM work_orders WHERE planned_start IS NOT NULL AND status != 'Pending' ORDER BY resource_id, planned_start").fetchall()
    else:
        orders = conn.execute("SELECT * FROM work_orders WHERE workshop=? AND planned_start IS NOT NULL AND status != 'Pending' ORDER BY resource_id, planned_start", (workshop,)).fetchall()
    conn.close()

    # 构建日期表头：从今天开始的 15 天
    #today = datetime.now().date()
    #date_list = [(today + timedelta(days=i)) for i in range(15)]
    #date_strs = [d.strftime('%Y-%m-%d') for d in date_list]

    # 构建日期表头：从明天开始的 15 天
    tomorrow = datetime.now().date() + timedelta(days=1)
    date_list = [(tomorrow + timedelta(days=i)) for i in range(15)]
    date_strs = [d.strftime('%Y-%m-%d') for d in date_list]

    matrix_data = []
    for row in orders:
        try:
            # 【修复核心】: 将 sqlite3.Row 转换为标准字典，防止 .get() 方法报错
            row_dict = dict(row) 
            
            # 解析时间
            start_dt = datetime.strptime(str(row_dict['planned_start'])[:16].replace('T', ' '), '%Y-%m-%d %H:%M')
            end_dt = datetime.strptime(str(row_dict['planned_end'])[:16].replace('T', ' '), '%Y-%m-%d %H:%M')

            # 计算总耗时 (分钟) 和总数量
            total_mins = max(1.0, (end_dt - start_dt).total_seconds() / 60.0)
            qty = float(row_dict['qty'] or 0)

            daily_dist = {}
            if qty > 0:
                d = start_dt.date()
                end_day = end_dt.date()

                # 将数量按天拆分
                while d <= end_day:
                    d_str = d.strftime('%Y-%m-%d')
                    day_start = datetime.combine(d, datetime.min.time())
                    day_end = day_start + timedelta(days=1)

                    # 计算当前日期的重叠时间
                    overlap_start = max(start_dt, day_start)
                    overlap_end = min(end_dt, day_end)
                    overlap_mins = max(0, (overlap_end - overlap_start).total_seconds() / 60.0)

                    if overlap_mins > 0:
                        # 按时间比例分配当天的排产数量
                        daily_qty = int(round(qty * (overlap_mins / total_mins)))
                        daily_dist[d_str] = daily_qty

                    d += timedelta(days=1)

            matrix_data.append({
                'task_id': row_dict['task_id'],
                'job_id': row_dict['job_id'],
                'product_code': row_dict['product_code'],
                'component_desc': row_dict['component_desc'] or '-',
                'resource_id': row_dict['resource_id'] or 'AUTO',
                'smt_side': row_dict.get('smt_side') or '-',
                'qty': int(qty),
                'planned_start': str(row_dict['planned_start'])[:16].replace('T', ' '),
                'status': row_dict['status'],
                'daily_dist': daily_dist
            })
        except Exception as e:
            import traceback
            traceback.print_exc() # 在后台打印具体的错误，防止死静
            continue

    return jsonify({
        'dates': date_strs,
        'orders': matrix_data
    })

# ====================================================
# === 新增：工厂日历管理 API (支持 CSV 导入导出) ===
# ====================================================
# ====================================================
# === 修复版：工厂日历管理 API (自带动态建表功能) ===
# ====================================================
def init_calendar_table(conn):
    """自动创建日历表（如果不存在的话）"""
    conn.execute('''
        CREATE TABLE IF NOT EXISTS factory_calendar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workshop_code TEXT,
            workshop_name TEXT,
            calendar_date TEXT,
            is_workday INTEGER,
            work_start_time TEXT,
            work_end_time TEXT,
            notes TEXT,
            UNIQUE(workshop_code, calendar_date)
        )
    ''')

@app.route('/api/calendar/list')
def get_calendar_list():
    """获取日历列表"""
    from database import get_db_connection
    workshop = request.args.get('workshop', 'SMT')
    
    conn = get_db_connection()
    init_calendar_table(conn)
    rows = conn.execute("SELECT * FROM factory_calendar WHERE workshop_code=? ORDER BY calendar_date", (workshop,)).fetchall()
    conn.close()
    
    return jsonify([dict(r) for r in rows])

@app.route('/api/calendar/generate', methods=['POST'])
def generate_calendar():
    """一键生成默认日历 (极致兼容版 + 详细Debug日志)"""
    from database import get_db_connection
    from datetime import datetime, timedelta
    from flask import request, jsonify
    import traceback
    
    print("\n=== [DEBUG] 开始处理 generate_calendar 请求 ===")
    print("Content-Type:", request.content_type)
    print("Raw Data (原始数据):", request.data)
    
    try:
        # 1. 安全获取参数
        data = request.get_json(force=True, silent=True) or {}
        print("Parsed JSON (解析后字典):", data)
        
        workshop = data.get('workshop', 'SMT')
        days = int(data.get('days', 30))
        
        ws_names = {'SMT': 'SMT车间', 'DIP': 'DIP车间', 'ASSEMBLY': '总装车间'}
        ws_name = ws_names.get(workshop, workshop)
        
        # 2. 连接数据库
        print(f"正在连接数据库，准备生成 {ws_name} 的 {days} 天日历...")
        conn = get_db_connection()
        init_calendar_table(conn)
        print("表结构初始化检查通过。")
        
        # 3. 生成数据
        start_date = datetime.now().date()
        for i in range(days):
            cur_date = start_date + timedelta(days=i)
            date_str = cur_date.strftime('%Y-%m-%d')
            is_weekend = cur_date.weekday() >= 5
            
            is_workday = 0 if is_weekend else 1
            notes = '周末休息' if is_weekend else '正常上班'
            
            # 【核心兼容性修复】：放弃 ON CONFLICT，改用先删后插，兼容所有旧版 SQLite
            conn.execute('DELETE FROM factory_calendar WHERE workshop_code=? AND calendar_date=?', (workshop, date_str))
            
            conn.execute('''
                INSERT INTO factory_calendar 
                (workshop_code, workshop_name, calendar_date, is_workday, work_start_time, work_end_time, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (workshop, ws_name, date_str, is_workday, '08:00', '20:00', notes))
            
        conn.commit()
        conn.close()
        print("=== [DEBUG] 数据生成并保存成功！ ===\n")
        return jsonify({'success': True, 'message': f'成功生成 {ws_name} 未来 {days} 天的日历'})
        
    except Exception as e:
        print("=== [DEBUG] 发生严重异常！ ===")
        traceback.print_exc() # 打印完整的报错堆栈到终端
        print("==============================\n")
        return jsonify({'success': False, 'message': f'后端报错: {str(e)}'})

# === 彻底防冲突的新导出路由 ===
@app.route('/api/calendar/download_csv')
def download_calendar_csv():
    """导出为 CSV (全新防冲突路由)"""
    from database import get_db_connection
    import csv, io
    from flask import Response, request
    
    workshop = request.args.get('workshop', 'SMT')
    conn = get_db_connection()
    
    # 确保表存在，防止报错
    conn.execute('''
        CREATE TABLE IF NOT EXISTS factory_calendar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workshop_code TEXT, workshop_name TEXT, calendar_date TEXT,
            is_workday INTEGER, work_start_time TEXT, work_end_time TEXT, notes TEXT,
            UNIQUE(workshop_code, calendar_date)
        )
    ''')
    
    rows = conn.execute("SELECT * FROM factory_calendar WHERE workshop_code=? ORDER BY calendar_date", (workshop,)).fetchall()
    conn.close()
    
    si = io.StringIO()
    cw = csv.writer(si)
    # 写入表头
    cw.writerow(['workshop_code', 'workshop_name', 'calendar_date', 'is_workday', 'work_start_time', 'work_end_time', 'notes'])
    for row in rows:
        cw.writerow([row['workshop_code'], row['workshop_name'], row['calendar_date'], 
                     row['is_workday'], row['work_start_time'], row['work_end_time'], row['notes']])
    
    # 加上 utf-8-sig (BOM头) 防止 Excel 打开中文乱码
    output = si.getvalue().encode('utf-8-sig') 
    return Response(output, mimetype='text/csv', headers={'Content-Disposition': f'attachment;filename=calendar_{workshop}.csv'})

# === 彻底防冲突的新导入路由 (强化 WPS/Excel 编码清洗 + 日期防篡改) ===
# === 带有强力 DEBUG 日志的 CSV 导入路由 ===
@app.route('/api/calendar/upload_csv', methods=['POST'])
def upload_calendar_csv():
    from database import get_db_connection
    import csv, io
    from flask import request, jsonify
    import traceback
    
    print("\n" + "="*50)
    print("=== [DEBUG] 开始处理 CSV 导入 ===")
    
    if 'file' not in request.files:
        print("[DEBUG] 失败：未找到上传的文件")
        return jsonify({'success': False, 'message': '未找到文件'})
        
    file = request.files['file']
    print(f"[DEBUG] 接收到文件: {file.filename}")
    
    try:
        raw_bytes = file.stream.read()
        print(f"[DEBUG] 文件大小: {len(raw_bytes)} bytes")
        
        # 1. 编码识别
        try:
            content = raw_bytes.decode("utf-8-sig")
            print("[DEBUG] 编码识别: 成功使用 utf-8-sig 解码")
        except UnicodeDecodeError:
            content = raw_bytes.decode("gbk", errors="ignore")
            print("[DEBUG] 编码识别: 成功使用 gbk 解码")

        # 打印文件前100个字符，看是否混入了奇怪的符号
        print(f"[DEBUG] 文件开头预览: {repr(content[:100])}")

        stream = io.StringIO(content.strip(), newline=None)
        reader = csv.DictReader(stream)
        
        print(f"[DEBUG] 解析到的表头 (Headers): {reader.fieldnames}")
        
        conn = get_db_connection()
        count = 0
        
        for i, row in enumerate(reader):
            # 暴力清洗字典里所有的 Key 和 Value
            clean_row = {str(k).strip('\ufeff \t"'): str(v).strip() for k, v in row.items() if k is not None}
            
            # 只打印前 3 行的数据，防止刷屏
            if i < 3:
                print("-" * 30)
                print(f"[DEBUG] 第 {i+1} 行 原始解析: {row}")
                print(f"[DEBUG] 第 {i+1} 行 清洗之后: {clean_row}")

            ws_code = clean_row.get('workshop_code')
            raw_date = clean_row.get('calendar_date', '')
            
            # 日期修复逻辑
            c_date = raw_date
            if raw_date:
                raw_date_temp = raw_date.replace('/', '-')
                parts = raw_date_temp.split('-')
                if len(parts) == 3:
                    try:
                        c_date = f"{parts[0]}-{int(parts[1]):02d}-{int(parts[2]):02d}"
                    except ValueError:
                        pass # 忽略非数字的转换错误
            
            if i < 3:
                print(f"[DEBUG] 第 {i+1} 行 日期转换: 原始 [{raw_date}] -> 最终 [{c_date}]")

            # 如果是空行则跳过
            if not ws_code or not c_date:
                if i < 3:
                    print(f"[DEBUG] 第 {i+1} 行 跳过: 缺少车间代码或日期为空")
                continue
                
            # 先删后插
            conn.execute('DELETE FROM factory_calendar WHERE workshop_code=? AND calendar_date=?', (ws_code, c_date))
            
            conn.execute('''
                INSERT INTO factory_calendar 
                (workshop_code, workshop_name, calendar_date, is_workday, work_start_time, work_end_time, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                ws_code, clean_row.get('workshop_name', ''), c_date,
                int(clean_row.get('is_workday', 1) or 1), 
                clean_row.get('work_start_time', '08:00'),
                clean_row.get('work_end_time', '20:00'), 
                clean_row.get('notes', '')
            ))
            count += 1
            
        conn.commit()
        conn.close()
        print(f"=== [DEBUG] 导入完成！成功处理了 {count} 条数据 ===")
        print("="*50 + "\n")
        
        return jsonify({'success': True, 'message': f'成功导入并更新了 {count} 天的日历数据！'})
        
    except Exception as e:
        print("=== [DEBUG] 发生严重异常 ===")
        traceback.print_exc()
        print("="*50 + "\n")
        return jsonify({'success': False, 'message': f'解析失败, 请检查文件内容: {str(e)}'})

# === 彻底防冲突的新路由 ===
@app.route('/api/calendar/build', methods=['POST'])
def build_calendar_new():
    from database import get_db_connection
    from datetime import datetime, timedelta
    from flask import request, jsonify
    import traceback
    
    try:
        # 绝对安全的无脑解析，无视任何 Header 格式
        raw_data = request.get_data(as_text=True)
        import json
        data = json.loads(raw_data) if raw_data else {}
        
        workshop = data.get('workshop', 'SMT')
        days = int(data.get('days', 30))
        ws_name = {'SMT': 'SMT车间', 'DIP': 'DIP车间', 'ASSEMBLY': '总装车间'}.get(workshop, workshop)
        
        conn = get_db_connection()
        init_calendar_table(conn)
        
        start_date = datetime.now().date()
        for i in range(days):
            cur_date = start_date + timedelta(days=i)
            date_str = cur_date.strftime('%Y-%m-%d')
            is_weekend = cur_date.weekday() >= 5
            
            is_workday = 0 if is_weekend else 1
            notes = '周末休息' if is_weekend else '正常上班'
            
            # 先删后插，无视 SQLite 版本限制
            conn.execute('DELETE FROM factory_calendar WHERE workshop_code=? AND calendar_date=?', (workshop, date_str))
            conn.execute('''
                INSERT INTO factory_calendar 
                (workshop_code, workshop_name, calendar_date, is_workday, work_start_time, work_end_time, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (workshop, ws_name, date_str, is_workday, '08:00', '20:00', notes))
            
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': f'成功生成 {ws_name} 未来 {days} 天的日历'})
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'生成失败: {str(e)}'})

# ====================================================
# === 新增：预排交期评估看板 API ===
# ====================================================

@app.route('/evaluation')
def evaluation_page():
    """渲染预排交期评估看板页面"""
    return render_template('evaluation.html')

@app.route('/api/evaluation/data')
def get_evaluation_data():
    """获取所有未排产工单，并进行交期与物料齐套预估"""
    from database import get_db_connection
    from datetime import datetime
    import random # 仅用于演示如果缺少 SRM 数据时的随机物料状态
    
    conn = get_db_connection()
    # 只评估处于 Pending (待排产) 状态的工单
    orders = conn.execute("SELECT * FROM work_orders WHERE status = 'Pending' ORDER BY deadline").fetchall()
    conn.close()
    
    results = []
    today = datetime.now().date()
    
    # 统计数据
    stats = {'total': len(orders), 'normal': 0, 'risk': 0, 'shortage': 0}
    
    for row in orders:
        order = dict(row)
        status = 'Normal'
        reason = '产能充裕，物料齐套'
        color = 'success'
        
        deadline_str = order.get('deadline')
        if deadline_str:
            try:
                # 兼容处理截取前10位 YYYY-MM-DD
                deadline_date = datetime.strptime(str(deadline_str)[:10], '%Y-%m-%d').date()
                days_left = (deadline_date - today).days
                
                # 模拟 SRM 物料齐套状态 (由于我们目前没接真实 SRM，用 hash 模拟部分缺料)
                # 真实场景中这里会读取 order.get('material_ready_level')
                is_shortage = (hash(order['task_id']) % 10) > 7 # 约 20% 概率模拟缺料
                
                if days_left < 0:
                    status = 'Shortage'
                    reason = f'已逾期 {abs(days_left)} 天！严重延误！'
                    color = 'danger'
                elif is_shortage and days_left < 6:
                    # T+6 内必须齐套，否则亮红灯
                    status = 'Shortage'
                    reason = f'交期仅剩 {days_left} 天，但 SRM 核心物料未齐套！'
                    color = 'danger'
                elif days_left <= 7:
                    # T+7 产能预警
                    status = 'Risk'
                    reason = f'交期较近 (T+{days_left})，请密切关注产能占用'
                    color = 'warning'
                else:
                    status = 'Normal'
                    reason = f'交期健康 (T+{days_left})，具备正常排产条件'
                    color = 'success'
                    
            except Exception as e:
                status = 'Risk'
                reason = '交期格式异常，无法准确评估'
                color = 'secondary'
                
        order['eval_status'] = status
        order['eval_reason'] = reason
        order['eval_color'] = color
        
        # 计入统计
        if status == 'Normal': stats['normal'] += 1
        elif status == 'Risk': stats['risk'] += 1
        elif status == 'Shortage': stats['shortage'] += 1
        
        results.append(order)
        
    return jsonify({'success': True, 'stats': stats, 'data': results})

# ====================================================
# === 新增：车间排产矩阵 Excel 一键导出 API ===
# ====================================================
@app.route('/api/export_schedule_excel/<workshop>')
def export_schedule_excel(workshop):
    """生成带有格式、冻结表头、每日明细的工业级 Excel 排产单"""
    import io
    from flask import send_file
    from database import get_db_connection
    from datetime import datetime, timedelta
    from urllib.parse import quote
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return "请先在终端运行 pip install openpyxl 安装 Excel 导出支持库", 500

    conn = get_db_connection()
    if workshop == 'ALL':
        orders = conn.execute("SELECT * FROM work_orders WHERE planned_start IS NOT NULL AND status != 'Pending' ORDER BY resource_id, planned_start").fetchall()
    else:
        orders = conn.execute("SELECT * FROM work_orders WHERE workshop=? AND planned_start IS NOT NULL AND status != 'Pending' ORDER BY resource_id, planned_start", (workshop,)).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{workshop} 车间排产单"

    # 构建日期表头：从今天开始的 15 天 (与前端矩阵保持一致)
    today = datetime.now().date()
    date_list = [(today + timedelta(days=i)) for i in range(15)]
    
    headers = ['分配产线', '工单号 (Task ID)', '产品型号', 'A/B面', '排产总数', '计划开始时间', '计划结束时间']
    date_headers = [d.strftime('%Y-%m-%d') for d in date_list]
    all_headers = headers + date_headers

    ws.append(all_headers)

    # 定义工业风 Excel 样式
    header_fill = PatternFill(start_color="1A2332", end_color="1A2332", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color="BFBFBF"), 
                         right=Side(style='thin', color="BFBFBF"), 
                         top=Side(style='thin', color="BFBFBF"), 
                         bottom=Side(style='thin', color="BFBFBF"))

    # 渲染表头并设置列宽
    for col_num, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
        col_letter = openpyxl.utils.get_column_letter(col_num)
        if col_num == 2: # 工单号
            ws.column_dimensions[col_letter].width = 28
        elif col_num in [6, 7]: # 完整时间
            ws.column_dimensions[col_letter].width = 18
        elif col_num > len(headers): # 日期列
            ws.column_dimensions[col_letter].width = 12
        else:
            ws.column_dimensions[col_letter].width = 14

    ws.row_dimensions[1].height = 25
    # 【极致体验】冻结前 7 列和第 1 行！向右滚动看日期时，订单信息永远吸附在左侧！
    ws.freeze_panes = 'H2' 

    # 填入数据与按天摊派计算
    for row in orders:
        row_dict = dict(row)
        try:
            start_dt = datetime.strptime(str(row_dict['planned_start'])[:16].replace('T', ' '), '%Y-%m-%d %H:%M')
            end_dt = datetime.strptime(str(row_dict['planned_end'])[:16].replace('T', ' '), '%Y-%m-%d %H:%M')
            
            total_mins = max(1.0, (end_dt - start_dt).total_seconds() / 60.0)
            qty = float(row_dict['qty'] or 0)
            
            daily_dist = {}
            if qty > 0:
                d = start_dt.date()
                end_day = end_dt.date()
                while d <= end_day:
                    d_str = d.strftime('%Y-%m-%d')
                    day_start = datetime.combine(d, datetime.min.time())
                    day_end = day_start + timedelta(days=1)
                    
                    overlap_start = max(start_dt, day_start)
                    overlap_end = min(end_dt, day_end)
                    overlap_mins = max(0, (overlap_end - overlap_start).total_seconds() / 60.0)
                    
                    if overlap_mins > 0:
                        daily_dist[d_str] = int(round(qty * (overlap_mins / total_mins)))
                    d += timedelta(days=1)

            # 组装基础行数据
            row_data = [
                row_dict['resource_id'] or '-',
                row_dict['task_id'],
                row_dict['product_code'],
                row_dict.get('smt_side') or '-',
                int(qty),
                str(row_dict['planned_start'])[:16],
                str(row_dict['planned_end'])[:16]
            ]
            
            # 填入按天摊派的数量
            for d_str in date_headers:
                row_data.append(daily_dist.get(d_str, ''))
                
            ws.append(row_data)
            
            # 渲染数据行样式
            current_row = ws.max_row
            for col_num in range(1, len(all_headers) + 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.alignment = center_align
                cell.border = thin_border
                
                # 如果这一天有排产任务，给单元格打上醒目的橙色底色 (呼应您前端的 warning-color)
                if col_num > len(headers) and cell.value != '':
                    cell.fill = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid") # 浅橙色
                    cell.font = Font(bold=True, color="B45309") # 深橙字

        except Exception as e:
            print(f"导出处理行出错: {e}")
            continue

    # 生成文件
    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)
    
    filename = f"排产下发单_{workshop}_{today.strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        excel_io,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=quote(filename)
    )

# ====================================================
# === 新增：厂长 KPI 决策大屏 API ===
# ====================================================
@app.route('/kpi')
def kpi_page():
    """渲染厂长 KPI 大屏页面"""
    return render_template('kpi.html')

@app.route('/api/kpi_data')
def get_kpi_data():
    """计算全局 OEE、按时交付率 (OTD) 与未来 15 天产能水位图"""
    from database import get_db_connection
    from datetime import datetime, timedelta
    
    conn = get_db_connection()
    # 提取所有已排产的有效工单
    orders = conn.execute("SELECT * FROM work_orders WHERE status != 'Pending' AND planned_start IS NOT NULL").fetchall()
    resources = conn.execute("SELECT * FROM resources").fetchall()
    conn.close()

    today = datetime.now().date()
    dates = [(today + timedelta(days=i)).strftime('%m/%d') for i in range(15)]
    
    # --- 1. 计算 OTD (按时交付率) ---
    total_scheduled = len(orders)
    on_time_count = 0
    delayed_count = 0
    
    for o in orders:
        if not o['deadline']:
            on_time_count += 1
            continue
        try:
            p_end = datetime.strptime(str(o['planned_end'])[:16].replace('T', ' '), '%Y-%m-%d %H:%M')
            dl = datetime.strptime(str(o['deadline'])[:10], '%Y-%m-%d')
            dl = datetime.combine(dl, datetime.max.time())
            if p_end <= dl:
                on_time_count += 1
            else:
                delayed_count += 1
        except:
            on_time_count += 1
            
    otd_rate = round(on_time_count / total_scheduled * 100, 1) if total_scheduled > 0 else 100

    # --- 2. 计算未来 15 天各车间的产能负荷率 (水位图) ---
    workshop_load_mins = {'SMT': [0]*15, 'DIP': [0]*15, 'ASSEMBLY': [0]*15}
    
    for o in orders:
        ws = o['workshop']
        if ws not in workshop_load_mins: continue
        try:
            start_dt = datetime.strptime(str(o['planned_start'])[:16].replace('T', ' '), '%Y-%m-%d %H:%M')
            end_dt = datetime.strptime(str(o['planned_end'])[:16].replace('T', ' '), '%Y-%m-%d %H:%M')
            
            d = start_dt.date()
            while d <= end_dt.date():
                day_idx = (d - today).days
                if 0 <= day_idx < 15:
                    day_start = datetime.combine(d, datetime.min.time())
                    day_end = day_start + timedelta(days=1)
                    overlap_start = max(start_dt, day_start)
                    overlap_end = min(end_dt, day_end)
                    overlap_mins = max(0, (overlap_end - overlap_start).total_seconds() / 60.0)
                    workshop_load_mins[ws][day_idx] += overlap_mins
                d += timedelta(days=1)
        except: pass

    # 统计每个车间的标准总产能 (假设每天单线满载排班 12 小时 = 720 分钟)
    ws_lines = {'SMT': 0, 'DIP': 0, 'ASSEMBLY': 0}
    for r in resources:
        ws = r.get('workshop')
        # 兼容总装车间的判断
        if ws in ws_lines: ws_lines[ws] += 1
        elif r['type'] == 'Production': ws_lines['ASSEMBLY'] += 1
            
    series_data = []
    # 转换为百分比负荷率，构建前端 ECharts 所需格式
    for ws, daily_mins in workshop_load_mins.items():
        max_capacity_mins = max(ws_lines[ws] * 720, 1) # 避免除 0 报错
        percent_loads = [round((m / max_capacity_mins) * 100, 1) for m in daily_mins]
        
        # 配色方案
        color = '#238636' if ws == 'SMT' else ('#1f6feb' if ws == 'DIP' else '#d97706')
        
        series_data.append({
            'name': f'{ws}车间',
            'type': 'line',
            'smooth': True,
            'symbol': 'circle',
            'symbolSize': 8,
            'itemStyle': {'color': color},
            'areaStyle': {
                'color': {
                    'type': 'linear', 'x': 0, 'y': 0, 'x2': 0, 'y2': 1,
                    'colorStops': [{'offset': 0, 'color': color}, {'offset': 1, 'color': 'transparent'}]
                },
                'opacity': 0.3
            },
            'data': percent_loads
        })
        
    return jsonify({
        'success': True,
        'otd_rate': otd_rate,
        'total_orders': total_scheduled,
        'delayed_count': delayed_count,
        'dates': dates,
        'load_series': series_data
    })

if __name__ == '__main__':
    app.run(debug=True, port=5000, host='0.0.0.0', threaded=True)